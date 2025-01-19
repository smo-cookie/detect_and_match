import re
import json
import openai
from pymongo import MongoClient
from docx import Document
from openpyxl import load_workbook

# 4.0버전으로 쓸 경우 유료
# api 키는 깃헙에 올릴 때는 지워야함
openai.api_key = ""

# MongoDB 설정
client = MongoClient("")
db = client["personal_info_db"]  # 유저 정보 저장
collection = db["detected_info"] # 탐지된 개인정보 저장장

# 정규표현식 패턴
patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    "주소": r"\b[가-힣]+시 [가-힣]+구 [가-힣]+동\b",
    "연락처": r"\b010-\d{4}-\d{4}\b",
    "생년월일": r"\b\d{4}[-/]\d{2}[-/]\d{2}\b",
    "계좌번호": r"\b\d{2,4}-\d{2,4}-\d{2,4}\b",
    "여권번호": r"\b[A-Z]{1}\d{8}\b",
    "이메일": r"\b[A-Za-z0-9._%+-]+@(?:[A-Za-z0-9-]+\.)+[A-Za-z]{2,}\b",
    "카드번호": r"\b\d{4}-\d{4}-\d{4}-\d{4}\b",
    "성명": r"\b[가-힣]{2,3}\b"
} # 추가가

# word
def extract_text_from_word(file_path):
    document = Document(file_path)
    return "\n".join([paragraph.text for paragraph in document.paragraphs])

# excel
def extract_text_from_excel(file_path):
    workbook = load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text

# 정규표현식으로 개인정보 탐지
def detect_pii_with_regex(content):
    results = {}
    for key, pattern in patterns.items():
        matches = re.findall(pattern, content)
        if matches:
            results[key] = matches
    return results

# api 탐지
def detect_sensitive_info_with_chatgpt(content, additional_info):  # content로 문서 내용을 받아 챗 지피티로 추가로 탐지
    prompt = f"""
    다음 텍스트에서 개인정보 및 추가 요청된 정보를 찾아주세요:
    - 개인정보에는 연락처, 이메일, 주민등록번호, 주소, 계좌번호 등 개인을 특정할 수 있는 정보가 포함됩니다.
    - 추가 요청 정보: {additional_info}
    
    반환 형식:
    {{
        "개인정보": {{
            "연락처": ["010-1234-5678", ...],
            "이메일": ["example@domain.com", ...],
            ...
        }},
        "추가 탐지 정보": {{
            "프로젝트 이름": ["Project Alpha", ...],
            "회사 이름": ["XYZ Corporation", ...],
            ...
        }}
    }}
    
    텍스트:
    {content}
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    try:
        return json.loads(response['choices'][0]['message']['content'])
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from ChatGPT"}

# db 저장 -> 정규표현식 결과, 챗 지피티 결과 따로 분리해서 저장
def save_to_mongodb(file_name, file_content, regex_results, chatgpt_results):
    document = {
        "file_name": file_name,
        "file_content": file_content,  # 전체 문서 내용
        "regex_results": regex_results,  # 정규표현식 탐지 결과
        "chatgpt_results": chatgpt_results  # ChatGPT 탐지 결과 -> 추가 탐지 결과
    }
    collection.insert_one(document)

def main(file_path, file_type, additional_info):
    # 문서 내용 추출
    if file_type == "word":
        content = extract_text_from_word(file_path)
    elif file_type == "excel":
        content = extract_text_from_excel(file_path)
    else:
        print("지원하지 않는 파일 형식입니다.")
        return

    # 개인정보 탐지
    regex_results = detect_pii_with_regex(content)
    chatgpt_results = detect_sensitive_info_with_chatgpt(content, additional_info)

    # db에 파일경로명, 파일 내용, 탐지 결과 저장
    save_to_mongodb(file_path, content, regex_results, chatgpt_results)

    # 디버깅
    print(json.dumps({
        "regex_results": regex_results,
        "chatgpt_results": chatgpt_results
    }, ensure_ascii=False, indent=4))

if __name__ == "__main__":
    import sys
    file_path = sys.argv[1]  # 파일 경로
    file_type = sys.argv[2]  # 파일 타입: "word" 또는 "excel"
    additional_info = sys.argv[3]  # 추가 탐지 요청 정보 -> 선택
    main(file_path, file_type, additional_info)
